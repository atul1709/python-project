import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

class EmployeeManagementSystem:
    def __init__(self, filename):
        self.filename = filename
        if not os.path.exists(self.filename):
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Employees"
            self.sheet.append(["ID", "Name", "Age", "Position", "Salary"])
            self.workbook.save(self.filename)
        else:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheet = self.workbook.active

    def add_employee(self, name, age, position, salary):
        last_row = self.sheet.max_row
        new_id = last_row
        self.sheet.append([new_id, name, age, position, salary])
        self.workbook.save(self.filename)
        print(f"Employee added successfully with ID: {new_id}")

    def view_all_employees(self):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            print(f"ID: {row[0]}, Name: {row[1]}, Age: {row[2]} Position: {row[3]}, Salary: {row[4]}")

    def update_employee(self, employee_id, name=None, age=None, position=None, salary=None):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == employee_id:
                if name:
                    row[1].value = name
                if age:
                    row[2].value = age
                if position:
                    row[3].value = position
                if salary:
                    row[4].value = salary
                self.workbook.save(self.filename)
                print(f"Employee with ID {employee_id} updated successfully")
                return
        print(f"Employee with ID {employee_id} not found")

    def delete_employee(self, employee_id):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == employee_id:
                self.sheet.delete_rows(row[0].row)
                self.workbook.save(self.filename)
                print(f"Employee with ID {employee_id} deleted successfully")
                return
        print(f"Employee with ID {employee_id} not found")

def main():
    ems = EmployeeManagementSystem("employees.xlsx")

    while True:
        print("\nEmployee Management System")
        print("1. Add Employee")
        print("2. View All Employees")
        print("3. Update Employee")
        print("4. Delete Employee")
        print("5. Exit")

        choice = input("Enter your choice (1-5): ")

        if choice == "1":
            name = input("Enter employee name: ")
            age =  int(input("Enter employee age: "))          
            position = input("Enter employee position: ")
            salary = float(input("Enter employee salary: "))
            ems.add_employee(name, age, position, salary)
        elif choice == "2":
            ems.view_all_employees()
        elif choice == "3":
            employee_id = int(input("Enter employee ID to update: "))
            name = input("Enter new name (press enter to skip): ")
            age =  input("Enter new age (press enter to skip): ")
            age =  int(age) if age else None
            position = input("Enter new position (press enter to skip): ")
            salary = input("Enter new salary (press enter to skip): ")
            salary = float(salary) if salary else None
            ems.update_employee(employee_id, name, age, position, salary)
        elif choice == "4":
            employee_id = int(input("Enter employee ID to delete: "))
            ems.delete_employee(employee_id)
        elif choice == "5":
            print("Thank you for using the Employee Management System. Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()